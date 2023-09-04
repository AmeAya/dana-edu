from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils import timezone, dateformat
from django.utils.translation import gettext_lazy as _
from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from .functions import *
from .models import *
from .serializers import *


def aboutUsView(request):
    context = {'urls': getUserUrls(request.user)}
    return render(request, 'about_us_page.html', context)


@login_required(login_url='login_url')
def addQuestionsView(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'GET':
        if 'variant' in request.session:
            context = {
                'variant': Variant.objects.get(pk=request.session['variant']),
                'subjects': Subject.objects.all(),
                'points': [1, 2],
                'urls': getUserUrls(request.user)
            }

            if 'question_number' in request.session.keys():
                context.update({'question_number': int(request.session['question_number'])})
            else:
                context.update({'question_number': 1})

            if 'selected_subject' in request.session.keys():
                context.update({'selected_subject': int(request.session['selected_subject'])})

            return render(request, 'add_questions_page.html', context)
        else:
            return redirect('add_questions_init_url')
    else:
        variant = Variant.objects.get(pk=request.POST.get('variant'))
        question_number = int(request.POST.get('question_number'))
        question_points = request.POST.get('question_points')
        subject = Subject.objects.get(pk=request.POST.get('subject'))
        question = Question(number=question_number, points=question_points, subject=subject, variant=variant)
        question_text = request.POST.get('question_text')
        if question_text:
            question.text = question_text.replace("\n", "\x85")
        if 'question_image' in request.FILES:
            question_image = request.FILES['question_image']
            question.image = question_image
        question.save()

        for i in range(1, int(request.POST.get('answers_count'))):
            is_correct = False
            is_empty = True
            if request.POST.get('is_correct_answer_' + str(i)):
                is_correct = True
            answer = Answer(is_correct=is_correct)
            answer_text = request.POST.get('answer_text_' + str(i))
            if 'answer_image_' + str(i) in request.FILES:
                answer_image = request.FILES['answer_image_' + str(i)]
                answer.image = answer_image
                is_empty = False
            if answer_text:
                answer.text = answer_text
                is_empty = False
            if not is_empty:
                answer.save()
                question.answers.add(answer)

        if 'question_number' in request.session.keys():
            request.session['question_number'] += 1
        else:
            request.session['question_number'] = int(request.POST.get('question_number')) + 1
        request.session['selected_subject'] = request.POST.get('subject')

        return redirect('add_questions_url')


@login_required(login_url='login_url')
def addQuestionsInitView(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'GET':
        context = {'variants': Variant.objects.all(), 'urls': getUserUrls(request.user)}
        return render(request, 'add_questions_init_page.html', context)
    else:
        request.session['variant'] = request.POST.get('variant')
        return redirect('add_questions_url')


@login_required(login_url='login_url')
def addVariantView(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'GET':
        from .functions import getExamTypesChoices
        exam_types = []
        for exam_type in getExamTypesChoices():
            exam_types.append(exam_type[0])
        langs = []
        for lang in getVariantLanguageChoices():
            langs.append(lang[0])
        context = {'exam_types': exam_types, 'langs': langs, 'urls': getUserUrls(request.user)}
        return render(request, 'add_variant_page.html', context)
    else:
        Variant(name=request.POST.get('variant_name'),
                exam_type=request.POST.get('exam_type'),
                language=request.POST.get('language')).save()
        return redirect('add_questions_init_url')


@login_required(login_url='login_url')
def endExamView(request):
    current_exam = CurrentExam.objects.get(user=request.user)
    starts_at = ExamForGroup.objects.filter(group=request.user.group).latest('pk').starts_at
    result = Result(user=request.user, variant=current_exam.variant, starts_at=starts_at, ends_at=timezone.now())
    pupil_answers = PupilAnswer.objects.filter(user=request.user).filter(is_in_action=True)
    result.points = 0
    result.save()
    for question in Question.objects.filter(variant=current_exam.variant):
        result.questions.add(question)
    for subject in current_exam.subjects.all():
        result.subjects.add(subject)
    result.save()

    for pupil_answer in pupil_answers:
        result.subjects.add(pupil_answer.question.subject)
        pupil_answer.points = getQuestionPoints(pupil_answer.question, pupil_answer.answers.all())
        result.answers.add(pupil_answer)
        result.points += pupil_answer.points
        result.save()
        pupil_answer.is_in_action = False
        pupil_answer.save()

    current_exam.delete()

    exam_group = ExamForGroup.objects.filter(group=request.user.group).latest('pk')
    exam_group.ended_users.add(request.user)
    exam_group.save()

    return redirect('exam_results_url')


@login_required(login_url='login_url')
def endExamViolatedView(request):
    if len(CurrentExam.objects.filter(user=request.user)) == 0:
        return redirect('exam_results_url')
    current_exam = CurrentExam.objects.get(user=request.user)
    starts_at = ExamForGroup.objects.filter(group=request.user.group).latest('pk').starts_at
    result = Result(user=request.user, variant=current_exam.variant, starts_at=starts_at, ends_at=timezone.now())
    result.points = 0
    result.is_violated = True
    result.save()

    current_exam.delete()

    exam_group = ExamForGroup.objects.filter(group=request.user.group).latest('pk')
    exam_group.ended_users.add(request.user)
    exam_group.save()

    return redirect('exam_results_url')


@login_required(login_url='login_url')
def examView(request):
    subjects = CurrentExam.objects.filter(user=request.user).latest('pk').subjects
    ends_at = ExamForGroup.objects.filter(group=request.user.group).latest('pk').ends_at
    context = {
        'subjects': subjects.all(),
        'ends_at': ends_at
    }
    return render(request, 'exam_page.html', context)


@login_required(login_url='login_url')
def examInitView(request):
    if request.user.type != 'PU':
        return redirect('home_url')
    group_exam = ExamForGroup.objects.filter(group=request.user.group)
    if group_exam:
        group_exam = group_exam.latest('pk')
        if group_exam.ends_at < timezone.now():
            context = {
                'message': _('Your exam already ends at ') + dateformat.format(group_exam.ends_at, 'Y-m-d H:i:s'),
                'urls': getUserUrls(request.user)
            }
            return render(request, 'exam_init_page.html', context)
        if request.user in group_exam.ended_users.all():
            context = {
                'message': _('You already done your exam!'),
                'urls': getUserUrls(request.user)
            }
            return render(request, 'exam_init_page.html', context)
    else:
        context = {
            'message': _('You don`t have Exams!'),
            'urls': getUserUrls(request.user)
        }
        return render(request, 'exam_init_page.html', context)
    current_exam = CurrentExam.objects.filter(user=request.user)
    if current_exam:
        if request.user.group.number == 4:
            context = {
                'message': _('Your variant is '),
                'variant': current_exam.latest('pk').variant,
                'urls': getUserUrls(request.user),
            }
            return render(request, 'exam_init_page.html', context)
        elif request.user.group.number == 9:
            context = {
                'message': _('Your variant is '),
                'variant': current_exam.latest('pk').variant,
                'urls': getUserUrls(request.user),
            }
            return render(request, 'exam_init_page.html', context)
        else:
            context = {
                'message': _('Your variant is '),
                'variant': current_exam.latest('pk').variant,
                'subjects': SubjectCombination.objects.all(),
                'urls': getUserUrls(request.user),
            }
            return render(request, 'exam_init_page.html', context)
    current_variant = getRandomVariant(group_exam.variants.all(), request.user)
    CurrentExam(user=request.user, variant=current_variant).save()
    context = {
        'message': _('Your variant is '),
        'variant': current_variant,
        'subjects': SubjectCombination.objects.all(),
        'urls': getUserUrls(request.user)
    }
    return render(request, 'exam_init_page.html', context)


@login_required(login_url='login_url')
def examResultsView(request):
    results = Result.objects.filter(user=request.user).order_by('-starts_at')

    context = {'results': results,
               'urls': getUserUrls(request.user)}
    return render(request, 'exam_results_page.html', context)


@login_required(login_url='login_url')
def examResultView(request, pk):
    result = Result.objects.get(pk=pk)
    pupil_answers = result.answers.all()
    if request.user.type != 'MO':
        if result.user != request.user:
            return redirect('exam_results_url')
    results = []
    for subject in result.subjects.all():
        results.append({'subject': subject, 'questions': []})
    for question in result.questions.all():
        correct_answers = []
        for answer in question.answers.all():
            if answer.is_correct:
                correct_answers.append(answer)
        this_question = {'question': question, 'answers': correct_answers, 'pupil_answers': [], 'points': 0}
        for pupil_answer in pupil_answers:
            if pupil_answer.question == question:
                this_question['pupil_answers'] = pupil_answer.answers.all()
                this_question['points'] = pupil_answer.points
                break
        if this_question['points'] == 1:
            this_question['points'] = _('1 point')
        else:
            this_question['points'] = str(this_question['points']) + _(' points')
        for subject in results:
            if subject['subject'] == question.subject:
                subject['questions'].append(this_question)
    context = {
        'is_violated': result.is_violated,
        'results': results,
        'variant': result.variant,
        'points': result.points,
        'starts_at': result.starts_at.strftime('%Y-%m-%d %H:%M'),
        'urls': getUserUrls(request.user)
    }
    return render(request, 'exam_result_page.html', context)


class GetAreaByRegion(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        region = Region.objects.get(id=request.POST.get('region_id'))
        area = Area.objects.filter(region=region)
        data = {'areas': AreaSerializer(area, many=True).data}
        return Response(data, status=status.HTTP_200_OK)


class GetExamQuestionApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        question = Question.objects.get(pk=request.GET.get('pk'))
        pupil_answer = PupilAnswer.objects.filter(user=request.user).filter(question=question).filter(is_in_action=True)
        answers = question.answers.all()
        data = {'question': QuestionSerializer(question).data,
                'answers': AnswerSerializer(answers, many=True).data}
        if pupil_answer:
            data.update({'pupil_answer': PupilAnswerSerializer(pupil_answer[0]).data})
        return Response(data, status=status.HTTP_200_OK)


class GetQuestionsBySubjectApiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        subject = Subject.objects.get(pk=request.query_params['subject_id'])
        variant = CurrentExam.objects.filter(user=request.user).latest('pk').variant
        questions = Question.objects.filter(subject=subject).filter(variant=variant).order_by('number')
        is_solved = []
        for question in questions:
            if PupilAnswer.objects.filter(user=request.user).filter(question=question).filter(is_in_action=True):
                is_solved.append(True)
            else:
                is_solved.append(False)
        data = {'questions': QuestionSerializer(questions, many=True).data,
                'is_solved': is_solved}
        return Response(data, status=status.HTTP_200_OK)


class GetSchoolByArea(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        area = Area.objects.get(id=request.POST.get('area_id'))
        school = School.objects.filter(area=area)
        data = {'schools': SchoolSerializer(school, many=True).data}
        return Response(data, status=status.HTTP_200_OK)


@login_required(login_url='login_url')
def getStatsView(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'GET':
        context = {
            'regions': Region.objects.all(),
            'urls': getUserUrls(request.user)
        }
        return render(request, 'get_stats_init_page.html', context)
    else:
        school = School.objects.get(id=request.POST.get('school'))
        groups = Group.objects.filter(school=school)
        results_by_groups = []
        for group in groups:
            results = []
            for user in CustomUser.objects.filter(group=group):
                results.append(Result.objects.filter(user=user))
            results_by_groups.append({group: results})
        context = {
            'results_by_groups': results_by_groups,
            'urls': getUserUrls(request.user),
            'school': school
        }
        return render(request, 'get_stats_page.html', context)


def homeView(request):
    if request.user.is_authenticated:
        context = {'urls': getUserUrls(request.user)}
    else:
        context = {'urls': getBaseUrls()}
    return render(request, 'home_page.html', context)


class ExcelStatsAPIVIew(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        school = School.objects.get(id=request.POST.get('school'))
        groups = Group.objects.filter(school=school)
        wb = Workbook()
        filename = school.name + '_stats.xlsx'
        for group in groups:
            ws = wb.create_sheet('new_sheet')
            ws.title = str(group.number) + group.literal
            pupils = CustomUser.objects.filter(group=group)
            for i in range(len(pupils)):
                student = ' '.join([pupils[i].surname, pupils[i].name, pupils[i].iin])
                ws.cell(row=i+2, column=1, value=student)
                results = Result.objects.filter(user=pupils[i])
                for j in range(len(results)):
                    temp = ' '.join([results[j].user.surname, results[j].user.name, results[j].user.iin])
                    ws.cell(row=i+2, column=1, value=temp)
                    ws.cell(row=1, column=j+2, value=results[j].starts_at)
                    ws.cell(row=i+2, column=j+2, value=results[j].points)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        wb.save('excel_stats/' + filename)
        with open('excel_stats/' + filename, 'rb') as file:
            response = HttpResponse(file, content_type='text/xlsx')
            response['Content-Disposition'] = 'attachment; filename=' + filename
            return response


@login_required(login_url='login_url')
def questionDeleteView(request, question_id):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'POST':
        question = Question.objects.get(id=question_id)
        request.session['deleted'] = {
            'variant': question.variant.pk,
            'subject': question.subject.pk
        }
        question.delete()
        return redirect('questions_update_url')


@login_required(login_url='login_url')
def questionUpdateView(request, question_id):
    if request.user.type != 'MO':
        return redirect('home_url')
    question = Question.objects.get(id=question_id)
    if request.method == 'GET':
        subjects = Subject.objects.all()
        answers = []
        counter = 1
        for answer in question.answers.all():
            answers.append({'number': counter, 'answer': answer})
            counter += 1
        context = {
            'question': question,
            'points': [1, 2],
            'subjects': subjects,
            'answers': answers,
            'answer_count': counter,
            'urls': getUserUrls(request.user)
        }
        return render(request, 'question_update_page.html', context)
    else:
        question = Question.objects.get(id=question_id)
        question.number = int(request.POST.get('question_number'))
        question.points = request.POST.get('question_points')
        question.subject = Subject.objects.get(pk=request.POST.get('subject'))
        question.text = request.POST.get('question_text')
        if 'question_image' in request.FILES:
            question_image = request.FILES['question_image']
            question.image = question_image
        question.save()

        for i in range(1, int(request.POST.get('answers_count'))):
            is_correct = False
            is_empty = True
            if request.POST.get('is_correct_answer_' + str(i)):
                is_correct = True
            if request.POST.get('answer_id_' + str(i)):
                answer = Answer.objects.get(id=request.POST.get('answer_id_' + str(i)))
            else:
                answer = Answer(is_correct=is_correct)
            answer_text = request.POST.get('answer_text_' + str(i))
            if 'answer_image_' + str(i) in request.FILES:
                answer_image = request.FILES['answer_image_' + str(i)]
                answer.image = answer_image
                is_empty = False
            if answer_text:
                answer.text = answer_text
                is_empty = False
            if not is_empty:
                answer.save()
                question.answers.add(answer)

        return redirect('question_update_url', question_id)


@login_required(login_url='login_url')
def questionsUpdateInitView(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if 'deleted' in request.session.keys():
        variant = Variant.objects.get(id=request.session['deleted']['variant'])
        subject = Subject.objects.get(id=request.session['deleted']['subject'])
        questions = Question.objects.filter(variant=variant).filter(subject=subject)
        context = {
            'questions': questions,
            'urls': getUserUrls(request.user)
        }
        return render(request, 'questions_update_page.html', context)
    if request.method == 'GET':
        variants = Variant.objects.all()
        subjects = Subject.objects.all()
        context = {
            'variants': variants,
            'subjects': subjects,
            'urls': getUserUrls(request.user)
        }
        return render(request, 'questions_update_init_page.html', context)
    else:
        variant = Variant.objects.get(id=request.POST.get('variant'))
        subject = Subject.objects.get(id=request.POST.get('subject'))
        questions = Question.objects.filter(variant=variant).filter(subject=subject)
        context = {
            'questions': questions,
            'urls': getUserUrls(request.user)
        }
        return render(request, 'questions_update_page.html', context)


@login_required(login_url='login_url')
def setExamForGroup(request):
    if request.user.type != 'MO':
        return redirect('home_url')
    if request.method == 'GET':
        context = {'groups': Group.objects.all(),
                   'vars_ru': Variant.objects.filter(language='RUS'),
                   'vars_kz': Variant.objects.filter(language='KAZ'),
                   'exam_for_groups': ExamForGroup.objects.all().order_by('-pk'),
                   'urls': getUserUrls(request.user)}
        return render(request, 'set_exam_for_groups_page.html', context)
    else:
        group = Group.objects.get(pk=int(request.POST.get('group')))
        start_time = request.POST.get('start_time')
        start_time = datetime.strptime(start_time, '%Y-%m-%dT%H:%M')
        duration = int(request.POST.get('duration'))
        exam_for_groups = ExamForGroup(group=group, duration=duration, starts_at=start_time)
        exam_for_groups.save()
        variants = request.POST.getlist('variants')
        for variant in variants:
            exam_for_groups.variants.add(Variant.objects.get(pk=int(variant)))
        exam_for_groups.save()
        return redirect('set_exam_for_groups_url')


class SetPupilAnswers(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        question = Question.objects.get(pk=request.POST.get('question'))
        pupil_answer = PupilAnswer.objects.filter(user=request.user).filter(question=question).filter(is_in_action=True)
        answers = Answer.objects.filter(pk__in=request.POST.get('answers').split(','))
        if not pupil_answer:
            pupil_answer = PupilAnswer(user=request.user, question=question)
            pupil_answer.save()
        else:
            pupil_answer = pupil_answer.latest('pk')
            pupil_answer.answers.clear()
        pupil_answer.answers.add(*answers)
        pupil_answer.save()
        return Response(status=status.HTTP_200_OK)


@login_required(login_url='login_url')
def subjectSelectView(request):
    if request.method == 'POST':
        current_exam = CurrentExam.objects.filter(user=request.user).latest('pk')
        variant = current_exam.variant
        subjects = set()
        for question in Question.objects.filter(variant=variant):
            subjects.add(question.subject)
        for subject in subjects:
            current_exam.subjects.add(subject)
        return redirect('exam_url')
    else:
        return redirect('home_url')


@login_required(login_url='login_url')
def userCabinetView(request):
    context = {'urls': getUserUrls(request.user)}
    avg = getUserAvg(request.user)
    context['avg'] = avg
    return render(request, 'cabinet_page.html', context)
